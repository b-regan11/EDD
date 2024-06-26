import os
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta
import pandas as pd
import sys

def create_table(filepath, filename, mach2_timeslot_table, mach5_timeslot_table, mach6_timeslot_table, mach9_timeslot_table):
    # Combine filepath and filename
    full_filename = os.path.join(filepath, filename)
    
    # Creating Excel writer object
    writer = pd.ExcelWriter(full_filename, engine='xlsxwriter')

    # Loop through the machines
    for m in range(1, 5):
        if m == 1:
            df = mach2_timeslot_table
        elif m == 2:
            df = mach5_timeslot_table
        elif m == 3:
            df = mach6_timeslot_table
        elif m == 4:
            df = mach9_timeslot_table

        # Convert 'Start Time' column to datetime
        df['Start Time'] = pd.to_datetime(df['Start Time'])

        # Extract unique dates and times
        unique_dates = df['Start Time'].dt.date.unique()
        unique_times = df['Start Time'].dt.time.unique()

        # Create a new DataFrame with 'BLANK' values
        blank_data = [['-----'] * len(unique_dates) for _ in range(len(unique_times))]
        blank_df = pd.DataFrame(blank_data, index=unique_times, columns=unique_dates)

        # Iterate over the original DataFrame and update the corresponding cell in blank_df
        for index, row in df.iterrows():
            date = row['Start Time'].date()
            time = row['Start Time'].time()
            job_number = row['Job Number']
            blank_df.loc[time, date] = job_number
        
        # Write DataFrame to a sheet in the Excel file
        if m == 1:
            sheet_name = f"Machine {2}"
        elif m == 2:
            sheet_name = f"Machine {5}"
        elif m == 3:
            sheet_name = f"Machine {6}"
        elif m == 4:
            sheet_name = f"Machine {9}"
        
        blank_df.to_excel(writer, sheet_name=sheet_name)
    
    # Close the Excel writer object to save the workbook
    writer.close()

def modify_workbook(filepath, filename, materials_table):
    # Load the Excel file
    full_filepath = os.path.join(filepath, filename)
    wb = load_workbook(full_filepath)

    # Define the colors
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    dark_blue_fill = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
    light_blue_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    dark_green_fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")
    light_green_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    dark_yellow_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="DDD9C4", end_color="DDD9C4", fill_type="solid")
    dark_red_fill = PatternFill(start_color="FB607F", end_color="FB607F", fill_type="solid")
    light_red_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    dark_purple_fill = PatternFill(start_color="DF73FF", end_color="DF73FF", fill_type="solid")
    light_purple_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")

    # Iterate over each sheet in the Excel file
    for sheet_name in wb.sheetnames:
        # Get the active sheet
        ws = wb[sheet_name]

        # Set the width of the first column to 18
        ws.column_dimensions['A'].width = 19

        # Iterate over the rows in the start time column (excluding the header)
        for row in range(2, ws.max_row + 1):
            # Get the cell value in the start time column
            start_time_cell = ws.cell(row=row, column=1)
            ws.cell(row=row, column=1).font = Font(bold=False)

            timeobj = datetime.strptime(start_time_cell.value, '%H:%M:%S')
            # if using MacOS
            if sys.platform == "darwin":
                first_value = timeobj.strftime('%-I:%M %p')
                second_value = (timeobj + timedelta(minutes=30)).strftime('%-I:%M %p')
                start_time_cell.value = first_value + " - " + second_value
            
            # if using Windows
            elif sys.platform == "win32":
                first_value = timeobj.strftime('%#I:%M %p')
                end_time = timeobj + timedelta(minutes=30)
                second_value = end_time.strftime('%#I:%M %p')
                start_time_cell.value = first_value + " - " + second_value


        # Set the width of the columns starting from the second column (B) to 17
        for col_num, column in enumerate(ws.iter_cols(min_col=2), start=2):
            ws.column_dimensions[column[0].column_letter].width = 17

        # Set the width of the last column to 17
        ws.column_dimensions[ws.cell(row=1, column=ws.max_column).column_letter].width = 17

        # Center the text of every cell in the sheet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Increase the font size for the first row of data
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)

        # Remove borders from all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border()

        # Set the value of the first cell (A1) to "Time" and bold the cell
        ws.cell(row=1, column=1).value = "Time"
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)

        # format each cell in the first row (excluding the first cell)
        for col in range(2, ws.max_column + 1):
            date_cell = ws.cell(row=1, column=col)            
            dateobj = date_cell.value
            if sys.platform == "darwin":
                date_value = dateobj.strftime('%a %-m/%-d/%y')
            elif sys.platform == "win32":
                date_value = dateobj.strftime('%a %#m/%#d/%y')
            date_cell.value = date_value

        # Change the fill color of cells
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=1).fill = grey_fill
        
        # Change the fill color of cells
        for col in range(2, ws.max_column + 1):  # Start from the second column
            # Check if the first row value starts with "Mon"
            if ws.cell(row=1, column=col).value.startswith("Mon"):
                # Apply light blue fill color to each cell in the column
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = light_blue_fill
                ws.cell(row=1, column=col).fill = dark_blue_fill
            
            elif ws.cell(row=1, column=col).value.startswith("Tue"):
                # Apply light green fill color to each cell in the column
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = light_green_fill
                ws.cell(row=1, column=col).fill = dark_green_fill
            
            elif ws.cell(row=1, column=col).value.startswith("Wed"):
                # Apply light yellow fill color to each cell in the column
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = light_yellow_fill
                ws.cell(row=1, column=col).fill = dark_yellow_fill
            
            elif ws.cell(row=1, column=col).value.startswith("Thu"):
                # Apply light red fill color to each cell in the column
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = light_red_fill
                ws.cell(row=1, column=col).fill = dark_red_fill
            
            elif ws.cell(row=1, column=col).value.startswith("Fri"):
                # Apply light purple fill color to each cell in the column
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = light_purple_fill
                ws.cell(row=1, column=col).fill = dark_purple_fill
            
            else:
                # Apply light purple fill color by default
                for row in range(1, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = grey_fill

        # Get the last row
        last_row = ws.max_row

        # Set borders for all filled cells
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                if cell_value is not None:  # Check if the cell is not empty
                    border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
                    # Apply the border to the cell
                    ws.cell(row=row_idx, column=col_idx).border = border

        # Add a row with the first cell copied from the cell above
        for col in range(1, ws.max_column + 1):
            ws.cell(row=last_row, column=col).value = ws.cell(row=last_row, column=col).value

    # Insert a row at the top of each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(row=1, column=1).value = sheet_name
        ws.cell(row=1, column=1).font = Font(bold=True, size=22)
        ws.cell(row=1, column=1).fill = grey_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        # Apply the border to all cells within the merged range
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thick'), 
                                    right=Side(style='thick'), 
                                    top=Side(style='thick'), 
                                    bottom=Side(style='thick'))
                
        DummyUsed = False

        # Iterate over the rows and columns in the sheet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Check if the cell value contains the word "dummy"
                if cell.value and "dummy" in str(cell.value).lower():
                    # Replace the cell value with "-----"
                    cell.value = "-----"
                    DummyUsed = True
        if DummyUsed == True:
            print("No more jobs were selectable from the input file")

    # Create a new sheet for the materials table
    ws_materials = wb.create_sheet(title='Materials')

    # Convert the materials table DataFrame to an Excel sheet
    for row in pd.DataFrame(materials_table).iterrows():
        ws_materials.append(row[1].tolist())
    ws_materials.insert_rows(1)
    ws_materials.cell(row=1, column=1).value = "Job_Num"
    ws_materials.cell(row=1, column=1).font = Font(bold=True)
    ws_materials.cell(row=1, column=2).value = "Lbs"
    ws_materials.cell(row=1, column=2).font = Font(bold=True)
    ws_materials.cell(row=1, column=3).value = "Material"
    ws_materials.cell(row=1, column=3).font = Font(bold=True)
    ws_materials.cell(row=1, column=4).value = "Cost per Pound"
    ws_materials.cell(row=1, column=4).font = Font(bold=True)
    ws_materials.cell(row=1, column=5).value = "Job Cost"
    ws_materials.cell(row=1, column=5).font = Font(bold=True)
    
    ws_materials.delete_cols(4)
    ws_materials.delete_cols(2)

    # Iterate through each column and set the width based on the maximum content length
    for column in ws_materials.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust the width (you can change the multiplier for better fit)
        ws_materials.column_dimensions[column_letter].width = adjusted_width

    # Write to Excel
    print(f"Excel file saved at: {os.path.abspath(filename)}")

    # Save the workbook
    wb.save(full_filepath)